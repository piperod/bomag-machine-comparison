#!/usr/bin/env python3
"""
extract_images.py
----------------

Extract images from Excel file and organize them into folders by sheet name.
"""

import os
import re
import hashlib
from pathlib import Path
import openpyxl
from openpyxl_image_loader import SheetImageLoader

def get_cell_coordinates(cell_ref: str) -> tuple[int, int]:
    """Convert Excel cell reference (e.g. 'A1') to row, col numbers."""
    match = re.match(r'([A-Z]+)(\d+)', cell_ref)
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    
    col_str, row_str = match.groups()
    row = int(row_str)
    
    # Convert column letters to number (A=1, B=2, etc.)
    col = 0
    for char in col_str:
        col = col * 26 + (ord(char.upper()) - ord('A') + 1)
    
    return row, col

def clean_filename(text: str, max_length: int = 50) -> str:
    """Clean text to create a valid filename, truncating if necessary."""
    # Remove invalid filename characters and replace spaces with underscores
    clean = re.sub(r'[^\w\s-]', '', text.lower())
    clean = re.sub(r'[-\s]+', '_', clean)
    
    # Truncate if too long, but keep the extension
    if len(clean) > max_length:
        clean = clean[:max_length]
    
    return clean

def extract_images(excel_path: str, output_dir: str = "assets") -> None:
    """Extract images from Excel file and save them to appropriate folders."""
    # Create output directory if it doesn't exist
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)

    # Load the workbook
    workbook = openpyxl.load_workbook(excel_path)
    
    # Process each sheet
    for sheet_name in workbook.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Create sheet-specific directory
        sheet_dir = output_path / sheet_name
        sheet_dir.mkdir(exist_ok=True)
        
        # Load images from sheet
        image_loader = SheetImageLoader(sheet)
        
        # Track processed images to avoid duplicates
        processed_images = set()
        
        # Get all images in the sheet
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if image_loader.image_in(cell.coordinate):
                    try:
                        # Get image data
                        image = image_loader.get(cell.coordinate)
                        if not image:
                            continue
                            
                        # Get machine name from the row (assuming it's in column A)
                        machine_name = sheet.cell(row=row, column=1).value
                        if not machine_name:
                            machine_name = f"image_{row}_{col}"
                            
                        # Clean up machine name for filename
                        base_filename = clean_filename(machine_name)
                        
                        # Add a hash of the cell coordinate to ensure uniqueness
                        cell_hash = hashlib.md5(cell.coordinate.encode()).hexdigest()[:8]
                        filename = f"{base_filename}_{cell_hash}.png"
                        
                        image_path = sheet_dir / filename
                        
                        # Check if we've already processed this image
                        if str(image_path) in processed_images:
                            continue
                            
                        # Save image
                        image.save(str(image_path))
                        processed_images.add(str(image_path))
                        
                        print(f"Saved image for {machine_name} to {image_path}")
                    except Exception as e:
                        print(f"Error processing cell {cell.coordinate}: {e}")

def main():
    """Main entry point."""
    excel_path = "assets/Bomag-comparison.xlsx"
    extract_images(excel_path)

if __name__ == "__main__":
    main() 